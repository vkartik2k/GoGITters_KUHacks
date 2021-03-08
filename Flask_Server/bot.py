from selenium import webdriver
# from selenium import webdriver
# excel file imports
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import getpass
from selenium.common.exceptions import NoSuchElementException


# time import
from time import sleep

def waiting_N_click(driver):
        try:
            driver.find_element_by_css_selector('div.uArJ5e.UQuaGc.Y5sE8d.uyXBBb.xKiqt').click()  # join now
            print("TRIED")
        except:
            # print("1")
            sleep(2)
            waiting_N_click(driver)


def bot_call(meet_link,password):
    gmail_id = "foc1803@gmail.com"
    gmail_password = password
    # meet_link = "https://meet.google.com/ips-ehtt-gcj"

    # Some necessary things for automation with google driver
    profile = webdriver.FirefoxProfile()
    # profile.set_preference('permissions.default.desktop-notification', 1)
    profile.set_preference('permissions.default_content_setting_values.media_stream_mic',1)
    profile.set_preference('permissions.default_content_setting_values.media_stream_camera',2)

    driver = webdriver.Firefox(firefox_profile=profile)

    # This part allows the notifications, mic and camera permissions
    # Pass the argument 1 to allow and 2 to block

    # Sign in to google
    # driver = webdriver.Firefox()
    # URL = 'https://www.premierleague.com/stats/top/clubs/wins?se=16'
    # driver.get(URL)
    driver.get('https://stackoverflow.com/users/signup?ssrc=head&returnurl=%2fusers%2fstory%2fcurrent%27')  # signing in to google through stack overflow
    sleep(2)
    driver.find_element_by_xpath('//*[@id="openid-buttons"]/button[1]').click()  # signing in with google
    driver.find_element_by_xpath('//input[@type="email"]').send_keys(gmail_id)  # entering the gmail id
    driver.find_element_by_xpath('//*[@id="identifierNext"]').click()
    sleep(2)
    driver.find_element_by_xpath('//input[@type="password"]').send_keys(gmail_password)  # entering the password
    driver.find_element_by_xpath('//*[@id="passwordNext"]').click()
    sleep(2)

    driver.get('https://meet.google.com/')

    # Enter the meeting
    # Case when logged in with personal gmail account
    driver.find_element_by_css_selector('input#i3').send_keys(meet_link)  # Enter a code or link
    sleep(1)
    driver.find_element_by_css_selector('button.VfPpkd-LgbsSe.VfPpkd-LgbsSe-OWXEXe-dgl2Hf.ksBjEc.lKxP2d.cjtUbb').click()  # join
    sleep(2)
    cam_mic_selectors = driver.find_elements_by_css_selector('div.U26fgb.JRY2Pb.mUbCce.kpROve')  # camera and mic
    for e in cam_mic_selectors:
        e.click()
    
    sleep(10)
    driver.find_element_by_css_selector('div.uArJ5e.UQuaGc.Y5sE8d.uyXBBb.xKiqt').click()
    link = None
    while not link:
        try:
            link = driver.find_element_by_css_selector('div.uArJ5e.UQuaGc.kCyAyd.QU4Gid.foXzLb.IeuGXd') # participant list
            print(link)
        except NoSuchElementException:
            sleep(5)
    link.click()

    # waiting_N_click(driver) # join now
    sleep(7)
    names = driver.find_elements_by_css_selector('span.ZjFb7c')
    print(names) # participants
    lis=[]
    for e in names[1:]:
        lis.append(e.text)
    return lis
    # print("NAMES PRINT BEFORE")
    # n = int(driver.find_element_by_css_selector('div.eUyZxf span.rua5Nb').text.strip('(').strip(')'))  # no. of participants present
    # print(n)
    # wb = load_workbook('Google_Attendance.xlsx')
    # sheet = wb['Attendance Sheet']

    # for name in names:
    #     for box in range(2, sheet.max_row+1):
    #         if name.text.lower() == (sheet.cell(row=box, column=column_index_from_string('A')).value.lower()):  # checks if the participant is present in the sheet
    #             sheet[f'B{box}'] = 'Present'  # marks present of that participant

    # wb.save('Google_Attendance.xlsx')
    # print(f'No. of participants : {n-1}')  # 1 participant adds up while taking attendance
    # print('Attendance taken successfully!')


    